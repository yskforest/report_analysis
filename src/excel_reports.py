from __future__ import annotations

import os
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from io_models import AnalysisInputs, TaskResult, _safe_num, _normalize_paths


def _build_hierarchy_agg(df: pd.DataFrame, path_col: str, max_level: int) -> pd.DataFrame:
    work = df.copy()
    work[path_col] = work[path_col].astype(str).str.replace("\\\\", "/", regex=False).str.strip("/")
    parts = work[path_col].str.split("/")
    depth = int(parts.map(len).max()) if len(work) else 0
    use_depth = min(depth, max_level)

    rows = [{
        "Level": 0,
        "Path": "ALL_FILES",
        "SourceLines": int(work["SourceLines"].sum()),
        "TotalLines": int(work["TotalLines"].sum()),
        "CommentLines": int(work["CommentLines"].sum()),
        "FileCount": len(work),
    }]

    for i in range(1, use_depth + 1):
        tmp = work.copy()
        tmp["Path"] = parts.map(lambda x: "/".join(x[:i]) if len(x) >= i else "").replace("", "root")
        agg = tmp.groupby("Path", dropna=False).agg(
            SourceLines=("SourceLines", "sum"),
            TotalLines=("TotalLines", "sum"),
            CommentLines=("CommentLines", "sum"),
            FileCount=("SourceLines", "size")
        ).reset_index()
        agg.insert(0, "Level", i)
        agg = agg.sort_values(by=["SourceLines", "TotalLines"], ascending=False)
        rows.extend(agg.to_dict("records"))

    return pd.DataFrame(rows)


def _split_path_levels(series: pd.Series, max_level: int = 15) -> pd.DataFrame:
    parts = series.astype(str).str.replace("\\\\", "/", regex=False).str.strip("/").str.split("/")
    return pd.DataFrame({
        f"Level{i}": parts.map(lambda x: x[i] if len(x) > i else "")
        for i in range(max_level + 1)
    }, index=series.index)


def _distribution(df: pd.DataFrame, col: str) -> pd.DataFrame:
    if col not in df.columns:
        return pd.DataFrame(columns=[col, "Count"])
    s = _safe_num(df[col]).round(0).astype(int)
    out = s.value_counts(dropna=False).sort_index().reset_index()
    out.columns = [col, "Count"]
    return out


def _apply_common_styles(ws: openpyxl.worksheet.worksheet.Worksheet, has_header: bool = True, freeze_panes: str | None = None, enable_filter: bool = True) -> None:
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    thin_border = Border(
        left=Side(style='thin', color='E0E0E0'),
        right=Side(style='thin', color='E0E0E0'),
        top=Side(style='thin', color='E0E0E0'),
        bottom=Side(style='thin', color='E0E0E0')
    )
    
    max_row = ws.max_row
    max_col = ws.max_column
    if max_row < 1 or max_col < 1:
        return

    # 1. ヘッダーの装飾と枠線
    if has_header:
        for c in range(1, max_col + 1):
            cell = ws.cell(row=1, column=c)
            cell.fill = header_fill
            cell.font = header_font
            cell.border = thin_border
            
        if enable_filter and max_row > 1:
            ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"

    # 2. データセルの数値フォーマットと枠線（iter_rows で高速に巡回、大容量データは枠線描画を制限）
    apply_border = (max_row <= 1000)
    for row in ws.iter_rows(min_row=2, max_row=max_row, max_col=max_col):
        for cell in row:
            if apply_border:
                cell.border = thin_border
            val = cell.value
            if isinstance(val, float):
                if val.is_integer():
                    cell.number_format = '#,##0'
                else:
                    cell.number_format = '#,##0.00'
            elif isinstance(val, int) and not isinstance(val, bool):
                cell.number_format = '#,##0'

    # 3. ウィンドウ枠固定
    if freeze_panes:
        ws.freeze_panes = freeze_panes
        
    # 4. 列幅の自動調整 (先頭50行サンプリング)
    for col_idx in range(1, max_col + 1):
        col_letter = get_column_letter(col_idx)
        max_len = 0
        for row in ws.iter_rows(min_row=1, max_row=min(max_row, 50), min_col=col_idx, max_col=col_idx):
            cell_val = str(row[0].value or '')
            max_len = max(max_len, len(cell_val))
        ws.column_dimensions[col_letter].width = max(min(max_len + 3, 50), 10)



def run_excel_report(inputs: AnalysisInputs) -> TaskResult:
    out_file = inputs.output_dir / "metrics_report.xlsx"
    try:
        und_file_csv = inputs.output_dir / "und" / "und_file.csv"
        und_func_csv = inputs.output_dir / "und" / "und_func.csv"
        cloc_filtered_csv = inputs.output_dir / "cloc" / "cloc_filtered.csv"
        pmd_clone_ratio_csv = inputs.output_dir / "pmd" / "pmd_clone_ratio.csv"
        git_diff_csv = inputs.output_dir / "git" / "git_diff_file_metrics.csv"
        metrics_merge_csv = inputs.output_dir / "metrics_merge.csv"
        
        # 追加データソース
        file_metrics_csv = inputs.output_dir / "file" / "file_metrics.csv"
        exceeded_funcs_csv = inputs.output_dir / "und" / "threshold_exceeded_functions.csv"
        exceeded_summary_csv = inputs.output_dir / "und" / "threshold_exceeded_summary.csv"
        exceeded_dir_summary_csv = inputs.output_dir / "und" / "threshold_exceeded_dir_summary.csv"
        
        if not (und_file_csv.exists() or und_func_csv.exists() or cloc_filtered_csv.exists() or 
                pmd_clone_ratio_csv.exists() or git_diff_csv.exists() or metrics_merge_csv.exists() or
                file_metrics_csv.exists()):
            return TaskResult(
                name="excel_report",
                executed=False,
                success=True,
                message="excel report skipped (no intermediate CSVs found)",
            )

        # 各種データの読み込み
        cloc_df = pd.read_csv(cloc_filtered_csv) if cloc_filtered_csv.exists() else None
        und_file_df = pd.read_csv(und_file_csv, dtype=object, na_filter=False) if und_file_csv.exists() else None
        und_func_df = pd.read_csv(und_func_csv, dtype=object, na_filter=False) if und_func_csv.exists() else None
        pmd_df = pd.read_csv(pmd_clone_ratio_csv) if pmd_clone_ratio_csv.exists() else None
        git_df = pd.read_csv(git_diff_csv) if git_diff_csv.exists() else None
        file_metrics_df = pd.read_csv(file_metrics_csv) if file_metrics_csv.exists() else None
        exceeded_funcs_df = pd.read_csv(exceeded_funcs_csv) if exceeded_funcs_csv.exists() else None
        exceeded_summary_df = pd.read_csv(exceeded_summary_csv) if exceeded_summary_csv.exists() else None
        exceeded_dir_summary_df = pd.read_csv(exceeded_dir_summary_csv) if exceeded_dir_summary_csv.exists() else None
        merge_df = pd.read_csv(metrics_merge_csv) if metrics_merge_csv.exists() else None

        # ── 1. サマリーデータの作成 ──
        summary_rows = []
        
        # 一般・規模メトリクス
        summary_rows.append({"項目 (Metric)": "■ 一般・規模メトリクス (General & Size)", "値 (Value)": ""})
        
        loc_val = 0
        if cloc_df is not None and not cloc_df.empty:
            loc_val = int(cloc_df["code"].sum())
        elif und_file_df is not None and not und_file_df.empty and "CountLineCode" in und_file_df.columns:
            loc_val = int(_safe_num(und_file_df["CountLineCode"]).sum())
        summary_rows.append({"項目 (Metric)": "  総ソースコード行数 (LoC)", "値 (Value)": loc_val})
        
        file_count = 0
        if file_metrics_df is not None and not file_metrics_df.empty:
            file_count = len(file_metrics_df)
        elif merge_df is not None and not merge_df.empty:
            file_count = len(merge_df)
        elif und_file_df is not None:
            file_count = len(und_file_df)
        summary_rows.append({"項目 (Metric)": "  総ファイル数", "値 (Value)": file_count})
        
        if file_metrics_df is not None and not file_metrics_df.empty:
            total_size_mb = float(_safe_num(file_metrics_df["file_size_bytes"]).sum()) / (1024 * 1024)
            summary_rows.append({"項目 (Metric)": "  総ファイルサイズ (MB)", "値 (Value)": round(total_size_mb, 2)})

        # 複雑度メトリクス
        if und_func_df is not None and not und_func_df.empty:
            summary_rows.append({"項目 (Metric)": "■ 複雑度メトリクス (Complexity)", "値 (Value)": ""})
            summary_rows.append({"項目 (Metric)": "  総関数・メソッド数", "値 (Value)": len(und_func_df)})
            if "Cyclomatic" in und_func_df.columns:
                cyc_nums = _safe_num(und_func_df["Cyclomatic"])
                summary_rows.append({"項目 (Metric)": "  平均サイクロマティック複雑度 (Avg Cyclomatic)", "値 (Value)": round(float(cyc_nums.mean()), 2)})
                summary_rows.append({"項目 (Metric)": "  最大サイクロマティック複雑度 (Max Cyclomatic)", "値 (Value)": int(cyc_nums.max()) if not cyc_nums.empty else 0})
            if "MaxNesting" in und_func_df.columns:
                nest_nums = _safe_num(und_func_df["MaxNesting"])
                summary_rows.append({"項目 (Metric)": "  最大ネスト数平均 (Avg MaxNesting)", "値 (Value)": round(float(nest_nums.mean()), 2)})
                summary_rows.append({"項目 (Metric)": "  プロジェクト最大ネスト数 (Max Nesting)", "値 (Value)": int(nest_nums.max()) if not nest_nums.empty else 0})
            if "Essential" in und_func_df.columns:
                summary_rows.append({"項目 (Metric)": "  基本複雑度平均 (Avg Essential)", "値 (Value)": round(float(_safe_num(und_func_df["Essential"]).mean()), 2)})

        # コード重複メトリクス
        if pmd_df is not None and not pmd_df.empty:
            summary_rows.append({"項目 (Metric)": "■ コード重複メトリクス (PMD Duplication)", "値 (Value)": ""})
            tot_tok = pmd_df["PmdTotalTokens"].sum()
            cln_tok = pmd_df["PmdCloneTokensSum"].sum()
            ratio = float(cln_tok / tot_tok * 100.0) if tot_tok else 0.0
            summary_rows.append({"項目 (Metric)": "  コード重複率 (Tokens %)", "値 (Value)": round(ratio, 2)})
            summary_rows.append({"項目 (Metric)": "  重複トークン数", "値 (Value)": int(cln_tok)})
            summary_rows.append({"項目 (Metric)": "  重複検知ファイル数", "値 (Value)": len(pmd_df[pmd_df["PmdCloneTokensSum"] > 0])})

        # Git 変更履歴メトリクス
        if git_df is not None and not git_df.empty:
            summary_rows.append({"項目 (Metric)": "■ Git 変更履歴メトリクス (Git Activity)", "値 (Value)": ""})
            summary_rows.append({"項目 (Metric)": "  累計追加行数 (Added Lines)", "値 (Value)": int(_safe_num(git_df["AddedLines"]).sum())})
            summary_rows.append({"項目 (Metric)": "  累計削除行数 (Deleted Lines)", "値 (Value)": int(_safe_num(git_df["DeletedLines"]).sum())})
            summary_rows.append({"項目 (Metric)": "  累計変更行数 (Changed Lines)", "値 (Value)": int(_safe_num(git_df["ChangedLines"]).sum())})

        # ファイルシステム特性
        if file_metrics_df is not None and not file_metrics_df.empty:
            summary_rows.append({"項目 (Metric)": "■ ファイルシステム特性 (File System)", "値 (Value)": ""})
            is_bin_str = file_metrics_df["is_binary"].astype(str).str.lower()
            text_files = len(file_metrics_df[is_bin_str == "false"])
            binary_files = len(file_metrics_df[is_bin_str == "true"])
            summary_rows.append({"項目 (Metric)": "  テキストファイル数", "値 (Value)": text_files})
            summary_rows.append({"項目 (Metric)": "  バイナリファイル数", "値 (Value)": binary_files})
            
            enc_series = file_metrics_df["encoding"].replace("", None).dropna()
            if not enc_series.empty:
                summary_rows.append({"項目 (Metric)": "  主要な文字コード (Dominant Encoding)", "値 (Value)": enc_series.mode().iloc[0]})
                non_utf8 = len(file_metrics_df[~file_metrics_df["encoding"].astype(str).str.lower().str.contains("utf-8|ascii", na=False)])
                if non_utf8 > 0:
                    summary_rows.append({"項目 (Metric)": "  非UTF-8 / 異物文字コードファイル数", "値 (Value)": non_utf8})
                
            le_series = file_metrics_df["line_ending"].replace(["N/A", ""], [None, None]).dropna()
            if not le_series.empty:
                summary_rows.append({"項目 (Metric)": "  主要な改行コード (Dominant Line Ending)", "値 (Value)": le_series.mode().iloc[0]})
                crlf_count = len(file_metrics_df[file_metrics_df["line_ending"].astype(str) == "CRLF"])
                if crlf_count > 0:
                    summary_rows.append({"項目 (Metric)": "  CRLF 改行コードファイル数", "値 (Value)": crlf_count})

        # 基準値超過メトリクス
        if exceeded_funcs_df is not None:
            summary_rows.append({"項目 (Metric)": "■ 基準値超過メトリクス (Threshold Violations)", "値 (Value)": ""})
            summary_rows.append({"項目 (Metric)": "  基準値超過関数・メソッド数", "値 (Value)": len(exceeded_funcs_df)})
            summary_rows.append({"項目 (Metric)": "  基準値超過ファイル数", "値 (Value)": len(exceeded_summary_df) if exceeded_summary_df is not None else 0})
            if exceeded_dir_summary_df is not None and not exceeded_dir_summary_df.empty:
                summary_rows.append({"項目 (Metric)": "  基準値超過ディレクトリ数", "値 (Value)": len(exceeded_dir_summary_df[exceeded_dir_summary_df["total_exceeded_count"] > 0])})
            if und_func_df is not None and len(und_func_df) > 0:
                violation_ratio = float(len(exceeded_funcs_df) / len(und_func_df) * 100.0)
                summary_rows.append({"項目 (Metric)": "  関数全体に対する超過比率", "値 (Value)": round(violation_ratio, 2)})

        summary_df = pd.DataFrame(summary_rows if summary_rows else [{"項目 (Metric)": "-", "値 (Value)": "-"}])

        with pd.ExcelWriter(out_file, engine="openpyxl") as writer:
            # 1. Summary シートを書き込む
            summary_df.to_excel(writer, sheet_name="summary", index=False)

            # 2. Hotspots シートを書き込む
            if merge_df is not None and not merge_df.empty:
                size_c = None
                for col in ["cloc_code", "und_CountLineCode", "und_CountLine"]:
                    if col in merge_df.columns:
                        size_c = col
                        break
                git_c = None
                for col in ["git_ChangedLines", "git_AddedLines"]:
                    if col in merge_df.columns:
                        git_c = col
                        break
                
                if size_c and git_c:
                    hot_df = merge_df.dropna(subset=["File", size_c, git_c]).copy()
                    hot_df[size_c] = pd.to_numeric(hot_df[size_c], errors="coerce").fillna(0)
                    hot_df[git_c] = pd.to_numeric(hot_df[git_c], errors="coerce").fillna(0)
                    hot_df = hot_df[(hot_df[size_c] > 0) & (hot_df[git_c] > 0)].copy()
                    
                    if not hot_df.empty:
                        s_min, s_max = hot_df[size_c].min(), hot_df[size_c].max()
                        g_min, g_max = hot_df[git_c].min(), hot_df[git_c].max()
                        
                        s_range = s_max - s_min
                        g_range = g_max - g_min
                        
                        s_norm = (hot_df[size_c] - s_min) / s_range if s_range > 0 else 1.0
                        g_norm = (hot_df[git_c] - g_min) / g_range if g_range > 0 else 1.0
                        
                        hot_df["HotspotScore"] = s_norm * g_norm
                        hot_df = hot_df.sort_values(by=["HotspotScore", git_c, size_c], ascending=False).head(20)
                        hot_df = _normalize_paths(hot_df, ["File"], inputs.remove_path_prefix)
                        
                        disp_df = pd.DataFrame({
                            "Rank": range(1, len(hot_df) + 1),
                            "File": hot_df["File"],
                            f"Code Size ({size_c})": hot_df[size_c],
                            f"Git Changes ({git_c})": hot_df[git_c],
                            "Hotspot Score": hot_df["HotspotScore"].round(4)
                        })
                        disp_df.to_excel(writer, sheet_name="hotspots", index=False)

            # 3. UND (File / Hierarchy)
            if und_file_csv.exists() and und_file_df is not None:
                max_level = int(os.getenv("FILE_METRICS_MAX_LEVEL", "8"))
                if max_level < 0:
                    max_level = 8
                
                required = {"File", "CountLineCode", "CountLineComment", "CountLine"}
                if required.issubset(und_file_df.columns):
                    metrics = pd.DataFrame({
                        "File": und_file_df["File"].astype(str),
                        "SourceLines": _safe_num(und_file_df["CountLineCode"]),
                        "CommentLines": _safe_num(und_file_df["CountLineComment"]),
                        "TotalLines": _safe_num(und_file_df["CountLine"]),
                    })
                    metrics = _normalize_paths(metrics, ["File"], inputs.remove_path_prefix)
                    metrics = metrics[metrics["File"].astype(str) != ""].copy()
                    metrics = metrics[
                        ~metrics["File"]
                        .astype(str)
                        .str.split("/")
                        .map(lambda parts: any(part.startswith(".") for part in parts if part))
                    ].copy()
                    
                    if not metrics.empty:
                        metrics = metrics.sort_values(by=["SourceLines", "TotalLines"], ascending=False)
                        metrics["Dir"] = metrics["File"].astype(str).map(lambda p: str(Path(p).parent).replace("\\", "/"))
                        metrics["Dir"] = metrics["Dir"].replace(".", "").str.strip("/")
                        
                        metrics_for_agg = metrics[metrics["Dir"] != ""].copy()
                        if not metrics_for_agg.empty:
                            hierarchy = _build_hierarchy_agg(metrics_for_agg, "Dir", max_level=max_level)
                            hierarchy["Path"] = hierarchy.apply(
                                lambda row: "  " * int(row["Level"]) + str(row["Path"]), axis=1
                            )
                            hierarchy.to_excel(writer, sheet_name="file_hierarchy", index=False)

            # 4. UND (Function / Detail / Agg / Dist)
            if und_func_csv.exists() and und_func_df is not None and not und_func_df.empty:
                und_func_df = _normalize_paths(und_func_df, ["File"], inputs.remove_path_prefix)
                und_func_df = und_func_df[und_func_df["File"].astype(str) != ""].copy()
                und_func_df = und_func_df[
                    ~und_func_df["File"].astype(str).str.split("/").map(lambda parts: any(part.startswith(".") for part in parts if part))
                ].copy()
                
                if not und_func_df.empty:
                    level_df = _split_path_levels(und_func_df["File"], max_level=15)
                    out_df = pd.concat([und_func_df.reset_index(drop=True), level_df.reset_index(drop=True)], axis=1)
                    for c in ["MaxNesting", "Cyclomatic", "Essential"]:
                        if c in out_df.columns:
                             out_df[c] = _safe_num(out_df[c])
                    
                    out_df.to_excel(writer, sheet_name="func_detail", index=False)
                    
                    agg_keys = [f"Level{i}" for i in range(16)]
                    agg_map = {"FunctionCount": ("File", "size")}
                    for metric in ["MaxNesting", "Cyclomatic", "Essential"]:
                        if metric in out_df.columns:
                            agg_map[f"{metric}_Avg"] = (metric, "mean")
                    
                    agg_df = out_df.groupby(agg_keys, dropna=False).agg(**agg_map).reset_index()
                    for c in agg_df.columns:
                        if c.endswith("_Avg"):
                            agg_df[c] = pd.to_numeric(agg_df[c], errors="coerce").round(3)
                    
                    agg_df.to_excel(writer, sheet_name="func_level_agg", index=False)
                    
                    dist_maxnesting = _distribution(out_df, "MaxNesting")
                    dist_cyclomatic = _distribution(out_df, "Cyclomatic")
                    dist_essential = _distribution(out_df, "Essential")
                    
                    dist_maxnesting.to_excel(writer, sheet_name="func_dist_nesting", index=False)
                    dist_cyclomatic.to_excel(writer, sheet_name="func_dist_cyclomatic", index=False)
                    dist_essential.to_excel(writer, sheet_name="func_dist_essential", index=False)

            # 5. CLOC
            if cloc_filtered_csv.exists() and cloc_df is not None and not cloc_df.empty:
                cloc_agg = cloc_df.groupby("language").agg(
                    code=("code", "sum"),
                    comment=("comment", "sum"),
                    blank=("blank", "sum"),
                    files=("code", "size")
                ).reset_index()
                cloc_agg = cloc_agg.sort_values(by="code", ascending=False)
                cloc_agg.to_excel(writer, sheet_name="cloc_summary", index=False)

            # 6. PMD
            if pmd_clone_ratio_csv.exists() and pmd_df is not None and not pmd_df.empty:
                pmd_df = _normalize_paths(pmd_df, ["File"], inputs.remove_path_prefix)
                pmd_df = pmd_df.sort_values(by="PmdCloneRatio", ascending=False)
                pmd_df.to_excel(writer, sheet_name="pmd_clone_ratio", index=False)

            # 7. Git Numstat
            if git_diff_csv.exists() and git_df is not None and not git_df.empty:
                git_df = _normalize_paths(git_df, ["File"], inputs.remove_path_prefix)
                git_df = git_df.sort_values(by="ChangedLines", ascending=False)
                git_df.to_excel(writer, sheet_name="git_diff", index=False)

            # 8. File Metrics
            if file_metrics_csv.exists() and file_metrics_df is not None and not file_metrics_df.empty:
                file_metrics_df = _normalize_paths(file_metrics_df, ["file"], inputs.remove_path_prefix)
                file_metrics_df = file_metrics_df.rename(columns={"file": "File"})
                file_metrics_df.to_excel(writer, sheet_name="file_metrics", index=False)

            # 9. Threshold Exceeded Summary / Detail
            if exceeded_dir_summary_csv.exists() and exceeded_dir_summary_df is not None and not exceeded_dir_summary_df.empty:
                exceeded_dir_summary_df = exceeded_dir_summary_df.sort_values(by="total_exceeded_count", ascending=False)
                exceeded_dir_summary_df.to_excel(writer, sheet_name="exceeded_dir_summary", index=False)

            if exceeded_summary_csv.exists() and exceeded_summary_df is not None and not exceeded_summary_df.empty:
                exceeded_summary_df = exceeded_summary_df.sort_values(by="total_exceeded_count", ascending=False)
                exceeded_summary_df = _normalize_paths(exceeded_summary_df, ["File"], inputs.remove_path_prefix)
                exceeded_summary_df.to_excel(writer, sheet_name="exceeded_summary", index=False)
                
            if exceeded_funcs_csv.exists() and exceeded_funcs_df is not None and not exceeded_funcs_df.empty:
                exceeded_funcs_df = _normalize_paths(exceeded_funcs_df, ["File"], inputs.remove_path_prefix)
                exceeded_funcs_df.to_excel(writer, sheet_name="exceeded_functions", index=False)


            # 10. Metrics Merge
            if metrics_merge_csv.exists() and merge_df is not None and not merge_df.empty:
                merge_df = _normalize_paths(merge_df, ["File"], inputs.remove_path_prefix)
                merge_df.to_excel(writer, sheet_name="metrics_merge", index=False)

            # ── 11. スタイル設定 & 枠線 & チャートの生成 ──
            for name, ws in writer.sheets.items():
                freeze = None
                if name in ["func_detail", "metrics_merge", "git_diff", "pmd_clone_ratio", "file_metrics", "exceeded_functions", "exceeded_summary", "exceeded_dir_summary", "hotspots"]:
                    freeze = "B2"
                elif name == "file_hierarchy":
                    freeze = "C2"
                    
                _apply_common_styles(ws, has_header=True, freeze_panes=freeze, enable_filter=True)
                
                # Summary シート用のカスタムスタイル
                if name == "summary":
                    ws.auto_filter.ref = None
                    for r in range(2, ws.max_row + 1):
                        cell_a = ws.cell(row=r, column=1)
                        cell_b = ws.cell(row=r, column=2)
                        val_a = str(cell_a.value or "")
                        if val_a.startswith("■"):
                            fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                            font = Font(bold=True, size=11, color="1F4E79")
                            cell_a.fill = fill
                            cell_b.fill = fill
                            cell_a.font = font
                            cell_b.font = font
                        else:
                            cell_a.font = Font(color="333333")
                            val_b = cell_b.value
                            if isinstance(val_b, (int, float)):
                                if "%" in val_a or "率" in val_a:
                                    cell_b.number_format = '0.00"%"'
                                elif isinstance(val_b, int) or (isinstance(val_b, float) and val_b.is_integer()):
                                    cell_b.number_format = '#,##0'
                                else:
                                    cell_b.number_format = '#,##0.00'

                # CLOC 円グラフの生成
                if name == "cloc_summary" and ws.max_row > 2:
                    try:
                        from openpyxl.chart import PieChart, Reference
                        pie = PieChart()
                        labels = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
                        data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
                        pie.add_data(data, titles_from_data=True)
                        pie.set_categories(labels)
                        pie.title = "Language Distribution (LoC)"
                        pie.width = 16
                        pie.height = 11
                        ws.add_chart(pie, "G2")
                    except Exception:
                        pass

                # 度数分布の棒グラフの生成
                if name in ["func_dist_nesting", "func_dist_cyclomatic", "func_dist_essential"] and ws.max_row > 2:
                    try:
                        from openpyxl.chart import BarChart, Reference
                        chart = BarChart()
                        chart.type = "col"
                        chart.style = 10
                        title_map = {
                            "func_dist_nesting": "Nesting Depth Distribution",
                            "func_dist_cyclomatic": "Cyclomatic Complexity Distribution",
                            "func_dist_essential": "Essential Complexity Distribution"
                        }
                        x_map = {
                            "func_dist_nesting": "Nesting Depth",
                            "func_dist_cyclomatic": "Cyclomatic Complexity",
                            "func_dist_essential": "Essential Complexity"
                        }
                        chart.title = title_map.get(name, "Distribution")
                        chart.y_axis.title = "Number of Functions"
                        chart.x_axis.title = x_map.get(name, "Metric")
                        
                        data = Reference(ws, min_col=2, min_row=1, max_row=ws.max_row)
                        cats = Reference(ws, min_col=1, min_row=2, max_row=ws.max_row)
                        chart.add_data(data, titles_from_data=True)
                        chart.set_categories(cats)
                        chart.legend = None
                        chart.width = 15
                        chart.height = 10
                        ws.add_chart(chart, "D2")
                    except Exception:
                        pass

                # 基準値超過トップ10ファイルの横棒グラフ
                if name == "exceeded_summary" and ws.max_row > 2:
                    try:
                        from openpyxl.chart import BarChart, Reference
                        chart = BarChart()
                        chart.type = "bar" # horizontal
                        chart.title = "Top 10 Files by Violation Count"
                        chart.x_axis.title = "File"
                        chart.y_axis.title = "Violation Count"
                        
                        data = Reference(ws, min_col=8, min_row=1, max_row=min(ws.max_row, 11))
                        cats = Reference(ws, min_col=1, min_row=2, max_row=min(ws.max_row, 11))
                        chart.add_data(data, titles_from_data=True)
                        chart.set_categories(cats)
                        chart.legend = None
                        chart.width = 18
                        chart.height = 12
                        ws.add_chart(chart, "J2")
                    except Exception:
                        pass
        
        return TaskResult(
            name="excel_report",
            executed=True,
            success=True,
            outputs=[out_file],
            message=f"unified excel report generated ({out_file.name})",
        )
    except Exception as exc:
        return TaskResult(
            name="excel_report",
            executed=True,
            success=False,
            message=f"excel report failed: {exc}",
        )